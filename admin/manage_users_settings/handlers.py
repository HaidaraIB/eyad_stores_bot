from telegram import Update, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes,
    CallbackQueryHandler,
    ConversationHandler,
    MessageHandler,
    filters,
)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
from datetime import datetime
from custom_filters import PrivateChatAndAdmin, PermissionFilter
from common.keyboards import (
    build_back_to_home_page_button,
    build_back_button,
    build_admin_keyboard,
)
from common.lang_dicts import TEXTS, get_lang
from common.common import format_datetime, format_float, escape_html
from admin.manage_users_settings.keyboards import (
    build_manage_users_settings_keyboard,
    build_user_balance_actions_keyboard,
)
from common.back_to_home_page import back_to_admin_home_page_handler
from start import admin_command, start_command
import tempfile
import os
import models
from decimal import Decimal


async def manage_users_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)
        keyboard = build_manage_users_settings_keyboard(lang)
        keyboard.append(build_back_button("back_to_admin_settings", lang=lang))
        keyboard.append(build_back_to_home_page_button(lang=lang, is_admin=True)[0])
        await update.callback_query.edit_message_text(
            text=TEXTS[lang]["manage_users_settings_title"],
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return ConversationHandler.END


manage_users_settings_handler = CallbackQueryHandler(
    manage_users_settings,
    "^manage_users_settings$|^back_to_manage_users_settings$",
)


async def export_users_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)

        await update.callback_query.answer(
            text=TEXTS[lang]["exporting_users"],
            show_alert=True,
        )

        await update.callback_query.delete_message()

        # إنشاء ملف Excel مؤقت
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            excel_path = tmp_file.name

            # إنشاء workbook جديد
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"

            # إضافة العناوين
            headers = [
                TEXTS[lang]["excel_user_id"],
                TEXTS[lang]["excel_username"],
                TEXTS[lang]["excel_name"],
                TEXTS[lang]["excel_language"],
                TEXTS[lang]["excel_is_admin"],
                TEXTS[lang]["excel_is_banned"],
                TEXTS[lang]["excel_created_at"],
            ]
            ws.append(headers)

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # جلب جميع المستخدمين من قاعدة البيانات
            with models.session_scope() as s:
                users = s.query(models.User).all()

                for user in users:
                    username = (
                        f"@{user.username}"
                        if user.username
                        else TEXTS[lang]["excel_no_username"]
                    )
                    lang_text = (
                        TEXTS[lang][f"lang_{user.lang.name.lower()}"]
                        if user.lang
                        else TEXTS[lang]["excel_unknown"]
                    )
                    is_admin = (
                        TEXTS[lang]["excel_yes"]
                        if user.is_admin
                        else TEXTS[lang]["excel_no"]
                    )
                    is_banned = (
                        TEXTS[lang]["excel_yes"]
                        if user.is_banned
                        else TEXTS[lang]["excel_no"]
                    )
                    created_at = (
                        format_datetime(user.created_at)
                        if user.created_at
                        else TEXTS[lang]["excel_unknown"]
                    )

                    ws.append(
                        [
                            user.user_id,
                            username,
                            user.name,
                            lang_text,
                            is_admin,
                            is_banned,
                            created_at,
                        ]
                    )

            # ضبط عرض الأعمدة
            ws.column_dimensions["A"].width = 15  # User ID
            ws.column_dimensions["B"].width = 20  # Username
            ws.column_dimensions["C"].width = 25  # Name
            ws.column_dimensions["D"].width = 15  # Language
            ws.column_dimensions["E"].width = 12  # Is Admin
            ws.column_dimensions["F"].width = 12  # Is Banned
            ws.column_dimensions["G"].width = 20  # Created At

            # حفظ الملف
            wb.save(excel_path)

        try:
            # إرسال الملف
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.xlsx"

            with open(excel_path, "rb") as excel_file:
                await context.bot.send_document(
                    chat_id=update.effective_user.id,
                    document=excel_file,
                    filename=filename,
                )

            text = TEXTS[lang]["users_exported_success"]
        except Exception as e:
            text = TEXTS[lang]["export_error"]

        # حذف الملف المؤقت
        if os.path.exists(excel_path):
            os.unlink(excel_path)

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=text + "\n\n" + TEXTS[lang]["continue_with_admin_command"],
        )


export_users_handler = CallbackQueryHandler(
    export_users_to_excel,
    "^export_users_to_excel$",
)

# Balance editing states
BALANCE_USER_ID, BALANCE_ACTION, BALANCE_AMOUNT = range(3)


async def edit_user_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Entry point for editing user balance"""
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(
            text=TEXTS[lang].get(
                "enter_user_id_for_balance",
                "أرسل معرف المستخدم الذي تريد تعديل رصيده:",
            ),
            reply_markup=InlineKeyboardMarkup(
                [
                    build_back_button("back_to_manage_users_settings", lang=lang),
                    build_back_to_home_page_button(lang=lang, is_admin=True)[0],
                ]
            ),
        )
        return BALANCE_USER_ID


async def get_user_id_for_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get user ID and show user info with balance action options"""
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)

        if update.message:
            user_id = int(update.message.text.strip())
            with models.session_scope() as s:
                user = s.get(models.User, user_id)
                if not user:
                    await update.message.reply_text(
                        text=TEXTS[lang]["user_not_found"],
                    )
                    return BALANCE_USER_ID
            context.user_data["balance_user_id"] = user_id
        else:
            user_id = context.user_data["balance_user_id"]

        with models.session_scope() as s:
            user = s.get(models.User, user_id)
            # Build user info text
            user_info = user.stringify(lang)
            balance_text = format_float(user.balance)

            text = f"<b>{TEXTS[lang].get('user_info', 'معلومات المستخدم')}:</b>\n\n"
            text += user_info
            text += f"\n\n{TEXTS[lang]['current_balance'].format(balance=balance_text)}"

            # Build keyboard with action options
            keyboard = build_user_balance_actions_keyboard(lang, user_id)
            keyboard.append(
                build_back_button("back_to_get_user_id_for_balance", lang=lang)
            )
            keyboard.append(build_back_to_home_page_button(lang=lang, is_admin=True)[0])
            if update.message:
                await update.message.reply_text(
                    text=text,
                    reply_markup=InlineKeyboardMarkup(keyboard),
                )
            else:
                await update.callback_query.edit_message_text(
                    text=text,
                    reply_markup=InlineKeyboardMarkup(keyboard),
                )
        return BALANCE_ACTION


back_to_get_user_id_for_balance = edit_user_balance


async def balance_action_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle balance action selection (add/deduct, set, zero)"""
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)
        data = update.callback_query.data

        # Extract action type and user_id
        if data.startswith("balance_action_add_deduct_"):
            action = "add_deduct"
            user_id = int(data.replace("balance_action_add_deduct_", ""))
        elif data.startswith("balance_action_set_"):
            action = "set"
            user_id = int(data.replace("balance_action_set_", ""))
        elif data.startswith("balance_action_zero_"):
            action = "zero"
            user_id = int(data.replace("balance_action_zero_", ""))
        else:
            return ConversationHandler.END

        context.user_data["balance_action"] = action
        context.user_data["balance_user_id"] = user_id

        if action == "zero":
            # Directly zero the balance
            with models.session_scope() as s:
                user = s.get(models.User, user_id)
                if user:
                    old_balance = user.balance
                    user.balance = Decimal("0.00")
                    s.commit()

                    balance_zeroed_text = (
                        TEXTS[lang]
                        .get(
                            "balance_zeroed",
                            f"تم تصفير الرصيد بنجاح ✅\nالرصيد السابق: {format_float(old_balance)} SDG",
                        )
                        .format(old_balance=format_float(old_balance))
                    )
                    await update.callback_query.answer(
                        text=balance_zeroed_text,
                        show_alert=True,
                    )
                    await update.callback_query.edit_message_text(
                        text=TEXTS[lang]["home_page"],
                        reply_markup=build_admin_keyboard(
                            lang, update.effective_user.id
                        ),
                    )
            return ConversationHandler.END
        else:
            # Ask for amount
            if action == "add_deduct":
                instruction = TEXTS[lang].get(
                    "enter_amount_add_deduct",
                    "أرسل المبلغ الذي تريد إضافته أو خصمه:\n(للسالب أرسل رقم سالب، مثال: -50)",
                )
            else:  # set
                instruction = TEXTS[lang].get(
                    "enter_new_balance",
                    "أرسل الرصيد الجديد:",
                )

            await update.callback_query.edit_message_text(
                text=instruction,
                reply_markup=InlineKeyboardMarkup(
                    [
                        build_back_button("back_to_balance_action_handler", lang=lang),
                        build_back_to_home_page_button(lang=lang, is_admin=True)[0],
                    ]
                ),
            )
            return BALANCE_AMOUNT


back_to_balance_action_handler = get_user_id_for_balance


async def process_balance_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Process the balance amount entered by admin"""
    if PrivateChatAndAdmin().filter(update) and PermissionFilter(
        models.Permission.MANAGE_USERS
    ).filter(update):
        lang = get_lang(update.effective_user.id)

        amount = Decimal(str(update.message.text.strip()))

        action = context.user_data.get("balance_action")
        user_id = context.user_data.get("balance_user_id")

        with models.session_scope() as s:
            user = s.get(models.User, user_id)

            old_balance = user.balance

            if action == "add_deduct":
                # Add or deduct amount
                user.balance += amount
                action_text_ar = "إضافة" if amount >= 0 else "خصم"
                action_text_en = "added" if amount >= 0 else "deducted"
                action_text = (
                    action_text_ar if lang == models.Language.ARABIC else action_text_en
                )
                result_text = (
                    TEXTS[lang]
                    .get(
                        "balance_updated_add_deduct",
                        f"تم {action_text} المبلغ بنجاح ✅\n"
                        f"الرصيد السابق: {format_float(old_balance)} SDG\n"
                        f"المبلغ: {format_float(abs(amount))} SDG\n"
                        f"الرصيد الجديد: {format_float(user.balance)} SDG",
                    )
                    .format(
                        action=action_text,
                        old_balance=format_float(old_balance),
                        amount=format_float(abs(amount)),
                        new_balance=format_float(user.balance),
                    )
                )
            else:  # set
                # Set new balance
                user.balance = amount
                result_text = (
                    TEXTS[lang]
                    .get(
                        "balance_updated_set",
                        f"تم تعيين الرصيد بنجاح ✅\n"
                        f"الرصيد السابق: {format_float(old_balance)} SDG\n"
                        f"الرصيد الجديد: {format_float(user.balance)} SDG",
                    )
                    .format(
                        old_balance=format_float(old_balance),
                        new_balance=format_float(user.balance),
                    )
                )

        await update.message.reply_text(
            text=result_text,
        )
        await update.message.reply_text(
            text=TEXTS[lang]["home_page"],
            reply_markup=build_admin_keyboard(lang, update.effective_user.id),
        )
        return ConversationHandler.END


edit_user_balance_handler = ConversationHandler(
    entry_points=[
        CallbackQueryHandler(
            edit_user_balance,
            "^edit_user_balance$",
        ),
    ],
    states={
        BALANCE_USER_ID: [
            MessageHandler(
                filters=filters.TEXT & ~filters.COMMAND,
                callback=get_user_id_for_balance,
            ),
        ],
        BALANCE_ACTION: [
            CallbackQueryHandler(
                balance_action_handler,
                r"^balance_action_(add_deduct|set|zero)_[0-9]+$",
            ),
        ],
        BALANCE_AMOUNT: [
            MessageHandler(
                filters=filters.TEXT & ~filters.COMMAND,
                callback=process_balance_amount,
            ),
        ],
    },
    fallbacks=[
        admin_command,
        start_command,
        back_to_admin_home_page_handler,
        manage_users_settings_handler,
        CallbackQueryHandler(
            back_to_get_user_id_for_balance,
            r"^back_to_get_user_id_for_balance$",
        ),
        CallbackQueryHandler(
            back_to_balance_action_handler,
            r"^back_to_balance_action_handler$",
        ),
    ],
)
